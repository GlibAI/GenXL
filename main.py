import io
import json
import os
import re

import openpyxl as xl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import google.genai as genai
from py_toon_format import encode
from dotenv import load_dotenv

load_dotenv()


def load_prompt(input_data: dict) -> str:
    """Build a prompt that converts raw extracted JSON into Excel cell mappings."""
    if not input_data:
        raise ValueError("Input data is empty")

    input_toon = encode(input_data)

    prompt = f"""You are a deterministic JSON-to-Excel layout engine with styling capabilities.

You receive extracted document data as JSON and must produce a structured JSON output
that maps every piece of data to exact Excel cell coordinates WITH styling properties.

────────────────────────
INPUT DATA (TOON encoded)
────────────────────────

{input_toon}

The input contains one or more files. Each file has:
- "file_name": name of the source document
- "classified_file_type": document category (e.g. "bank_statement")
- "fields": list of extracted fields, each with:
    - "field_name": human-readable label
    - "field_key": machine identifier
    - "section": logical grouping (e.g. "Account Information", "Transaction History")
    - "data_type": one of "String", "Number", "Date", "Table"
    - "value": the extracted value (can be null, a scalar, or an array of objects for Table types)

────────────────────────
YOUR TASK
────────────────────────

Produce a JSON object where:
- Each root key is an Excel SHEET NAME (derived from the file's classified_file_type, e.g. "Bank Statement").
- Each sheet value is a list of cell objects with the structure:
    {{
        "cell_coordinate": "<column_letter><row_number>",
        "cell_value": "<string or number>",
        "font_size": <integer>,
        "font_color": "<6-char hex color>",
        "background_color": "<6-char hex color or null>",
        "is_bold": <boolean>,
        "is_italic": <boolean>,
        "horizontal_alignment": "<left|center|right>",
        "vertical_alignment": "<top|center|bottom>",
        "border_top": "<thin|medium|thick|none>",
        "border_bottom": "<thin|medium|thick|none>",
        "border_left": "<thin|medium|thick|none>",
        "border_right": "<thin|medium|thick|none>",
        "border_color": "<6-char hex color>"
    }}

────────────────────────
LAYOUT RULES
────────────────────────

1. DOCUMENT HEADER (Row 1)
   - A1: The document title derived from classified_file_type (e.g. "Bank Statement")

2. SCALAR FIELDS (key-value pairs)
   - Group scalar fields (String, Number, Date) by their "section".
   - For each section, output a SECTION HEADER row with the section name in column A.
   - Then output each field as a row with:
       - Column A: field_name (the label)
       - Column B: the value (use "" if value is null)
   - Leave one empty row between sections for readability.

3. TABLE FIELDS (data_type = "Table")
   - Start on a new row after all scalar sections, with one empty row gap.
   - Output a SECTION HEADER row with the section name in column A.
   - Next row: column headers derived from the table's object keys (e.g. "Transaction Date", "Description", "Deposit Amount", etc.).
       - Place headers starting from column A, one per column.
   - Subsequent rows: one row per table record, values placed in corresponding columns.
   - Use "" for any null values within table rows.
   - Preserve the original column order from the input data.

4. COORDINATE ASSIGNMENT
   - Row numbering is sequential starting from 1.
   - Column letters follow Excel convention: A, B, C, ... Z, AA, AB, etc.
   - Every cell_coordinate must be unique within a sheet.
   - No gaps in row numbering except the intentional empty rows between sections.

────────────────────────
STYLING RULES (PRIORITY-BASED)
────────────────────────

You MUST assign styling to every cell based on its role. Use the hierarchy below
to decide font size, colors, weight, and alignment. Pick a professional, clean
color palette (blues/grays work well for financial documents).

PRIORITY 1 — DOCUMENT TITLE (highest visual priority)
   - font_size: 12
   - font_color: "000000"
   - background_color: "FFD2BF"
   - is_bold: true
   - is_italic: false
   - horizontal_alignment: "left"
   - vertical_alignment: "center"
   - border_top: "medium", border_bottom: "medium", border_left: "medium", border_right: "medium"
   - border_color: "000000"

PRIORITY 2 — SECTION HEADERS (section name rows)
   - font_size: 11
   - font_color: "000000"
   - background_color: "B6C2DB"
   - is_bold: true
   - is_italic: false
   - horizontal_alignment: "left"
   - vertical_alignment: "center"
   - border_top: "medium", border_bottom: "medium", border_left: "thin", border_right: "thin"
   - border_color: "000000"

PRIORITY 3 — FIELD LABELS (column A in key-value rows) & TABLE COLUMN HEADERS
   - font_size: 10
   - font_color: "000000"
   - background_color: "F0EFE8"
   - is_bold: true
   - is_italic: false
   - horizontal_alignment: "left"
   - vertical_alignment: "center"
   - border_top: "thin", border_bottom: "thin", border_left: "thin", border_right: "thin"
   - border_color: "000000"

PRIORITY 4 — FIELD VALUES (column B in key-value rows)
   - font_size: 10
   - font_color: "000000"
   - background_color: null (no fill)
   - is_bold: false
   - is_italic: false
   - horizontal_alignment: "left" for strings/dates, "right" for numbers
   - vertical_alignment: "center"
   - border_top: "thin", border_bottom: "thin", border_left: "thin", border_right: "thin"
   - border_color: "D3D3D3"

PRIORITY 5 — TABLE DATA CELLS (lowest priority)
   - font_size: 10
   - font_color: "000000"
   - background_color: null (no fill)
   - is_bold: false
   - is_italic: false
   - horizontal_alignment: "left" for text/dates, "right" for numeric columns
   - vertical_alignment: "center"
   - border_top: "thin", border_bottom: "thin", border_left: "thin", border_right: "thin"
   - border_color: "D3D3D3"

ADDITIONAL STYLING NOTES:
- For Number and currency data_type values, use "right" horizontal_alignment.
- For Date data_type values, use "center" horizontal_alignment.
- Use consistent colors across the entire sheet — do NOT vary colors per row.
- All hex colors must be exactly 6 characters (no # prefix).
- border_top/bottom/left/right must be one of: "thin", "medium", "thick", or "none".
- border_color is a single 6-char hex color applied to all four sides of the cell.
- Every cell MUST include all 5 border fields (border_top, border_bottom, border_left, border_right, border_color).

────────────────────────
OUTPUT REQUIREMENTS
────────────────────────

- Output MUST be raw, valid JSON only.
- DO NOT wrap in ```json or any code fences.
- DO NOT include markdown, comments, or explanations.
- The first character MUST be "{{" and the last MUST be "}}".
- Sheet names must NOT contain special characters (use only alphanumeric and spaces).
- cell_value must be a string or number — no nested objects.
- Numeric values should remain as numbers (not stringified).
- Date values should remain as strings in their original format.
- Every cell object MUST include all 12 styling fields (7 original + 5 border fields).
- background_color can be null for cells with no fill.
- Output must be deterministic and repeatable.
- Do NOT invent or fabricate data that is not present in the input.
- If a field's value is null, use an empty string "".
"""

    return prompt


def parse_llm_json(raw_text: str) -> dict:
    """Extract JSON from LLM output, handling code fences and stray text."""
    text = raw_text.strip()

    if text.startswith("```"):
        text = re.sub(r"^```[a-zA-Z]*", "", text)
        text = re.sub(r"```$", "", text)
        text = text.strip()

    first_brace = text.find("{")
    last_brace = text.rfind("}")

    if first_brace == -1 or last_brace == -1:
        raise ValueError("No JSON object found in LLM response")

    text = text[first_brace : last_brace + 1]
    return json.loads(text)


def call_llm(prompt: str) -> dict:
    """Call Gemini API and return the parsed response."""
    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        raise ValueError("GEMINI_API_KEY environment variable is not set")

    client = genai.Client(api_key=api_key)
    response = client.models.generate_content(
        model="gemini-2.5-flash",
        contents=prompt,
    )

    return parse_llm_json(response.text)


def apply_cell_style(cell, style_data: dict) -> None:
    """Apply styling properties from LLM output to an openpyxl cell."""
    font_size = style_data.get("font_size", 11)
    font_color = style_data.get("font_color", "000000")
    is_bold = style_data.get("is_bold", False)
    is_italic = style_data.get("is_italic", False)
    bg_color = style_data.get("background_color")
    h_align = style_data.get("horizontal_alignment", "left")
    v_align = style_data.get("vertical_alignment", "center")

    cell.font = Font(
        size=font_size,
        color=font_color,
        bold=is_bold,
        italic=is_italic,
    )

    if bg_color:
        cell.fill = PatternFill(
            start_color=bg_color,
            end_color=bg_color,
            fill_type="solid",
        )

    cell.alignment = Alignment(
        horizontal=h_align,
        vertical=v_align,
        wrap_text=True,
    )

    border_color = style_data.get("border_color", "D3D3D3")
    border_top = style_data.get("border_top", "none")
    border_bottom = style_data.get("border_bottom", "none")
    border_left = style_data.get("border_left", "none")
    border_right = style_data.get("border_right", "none")

    def make_side(style):
        if style and style != "none":
            return Side(style=style, color=border_color)
        return Side(style=None)

    cell.border = Border(
        top=make_side(border_top),
        bottom=make_side(border_bottom),
        left=make_side(border_left),
        right=make_side(border_right),
    )


def generate_excel(output_path: str, output_data_mappings: dict) -> None:
    """Apply mappings and styles to a new Excel workbook and save."""
    if not output_data_mappings:
        raise ValueError("Output data mappings are empty")

    final_workbook = xl.Workbook()

    for sheet_name, values in output_data_mappings.items():
        worksheet = final_workbook.create_sheet(sheet_name)

        for value in values:
            cell_coordinate = value.get("cell_coordinate")
            cell_value = value.get("cell_value")

            if not cell_coordinate:
                continue

            cell = worksheet[cell_coordinate]
            cell.value = cell_value
            apply_cell_style(cell, value)

    if "Sheet" in final_workbook.sheetnames:
        del final_workbook["Sheet"]

    buf = io.BytesIO()
    final_workbook.save(buf)
    buf.seek(0)

    with open(output_path, "wb") as f:
        f.write(buf.read())


def main():
    """Main execution flow: JSON input -> LLM -> structured JSON -> Excel."""
    input_json_path = "testing_jsons/testing_json_1.json"
    output_path = "output.xlsx"

    if not os.path.exists(input_json_path):
        raise FileNotFoundError(f"Input file not found: {input_json_path}")

    with open(input_json_path, "r", encoding="utf-8") as f:
        input_data = json.load(f)

    prompt = load_prompt(input_data)
    output_data_mappings = call_llm(prompt)
    generate_excel(output_path, output_data_mappings)

    print(f"Excel generated successfully: {output_path}")


if __name__ == "__main__":
    main()
